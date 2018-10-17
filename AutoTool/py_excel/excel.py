#!/usr/bin/env python
# coding:utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import sys


def pause():
    print '\n'
    print '\n'
    print '--------------------Work over-----------------------------'
    raw_input(" Please press any key to exit.")
    sys.exit(0)


class OperateExcel:

    def __init__(self):
        return

    # 导入excel文件中的数据
    def load_excel(self, load_filename):
        print("loading Excel")
        # 打开一个workbook
        if not os.path.exists(load_filename):
            print "No corresponding Excel file found."
            pause()

        try:
            wb = load_workbook(load_filename)
        except Exception, e:
            print "Excel is wrong : ", e
            pause()
        else:
            # print(wb.sheetnames)
            sheetnames = wb.sheetnames
            first_sheet = wb[sheetnames[0]]

            print "Work Sheet Title : ", first_sheet.title
            # print "Work Sheet Max Rows:", first_sheet.max_row
            # print "Work Sheet Min Rows:", first_sheet.min_row
            # print "Work Sheet Max Cols:", first_sheet.max_column
            # print "Work Sheet Min Cols:", first_sheet.min_column

            post_data = []
            headers = []
            for row in range(first_sheet.min_row, first_sheet.max_row + 1):
                post_content = {}
                for column in range(first_sheet.min_column, first_sheet.max_column + 1):
                    if row == 1:
                        value = first_sheet.cell(row=row, column=column).value
                        headers.append(value.lower().replace(' ', '').replace('\n', ''))
                    else:
                        if column == 1 and not first_sheet.cell(row=row, column=column).value:
                            continue
                        else:
                            post_content[headers[column - 1]] = first_sheet.cell(row=row, column=column).value

                if post_content and post_content:
                    post_data.append(post_content)

            # print "excel : ", post_data
            return post_data

    # 将数据导出到excel文件中
    def export_excel(self, store_filename, datas=None):
        print("export_excel")
        if os.path.exists(store_filename):
            os.remove(store_filename)

        wb = Workbook()
        first_sheet = wb.active
        first_sheet.title = u"sheet1"
        if datas is not None:
            # print datas
            i = 1
            j = 1
            for data in datas:
                for key in data:
                    print 'key is ', key
                    first_sheet.cell(row=i, column=j, value=key)
                    for value in data[key]:
                        i += 1
                        print value
                        first_sheet.cell(row=i, column=j, value=value)
                i = 1
                j += 1
        wb.save(store_filename)
        print "Excel save success"

        # 将数据导出到excel文件中

    def export_username_excel(self, store_filename, datas=None):
        print("export username to excel")
        if os.path.exists(store_filename):
            os.remove(store_filename)

        wb = Workbook()
        first_sheet = wb.active
        first_sheet.title = u"sheet1"
        first_sheet.cell(row=1, column=1, value="username")
        # print "\n"
        if datas is not None:
            # print datas
            i = 2
            j = 1
            for data in datas:
                # print 'username is ', data
                first_sheet.cell(row=i, column=j, value=data)
                i += 1
        wb.save(store_filename)
        print "Excel save success"

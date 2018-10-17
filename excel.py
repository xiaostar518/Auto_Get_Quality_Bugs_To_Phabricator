#!/usr/bin/env python
# coding:utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import util
import sys

columnName = util.columnName
priority = util.priority
title = util.title


# 将数据导出到excel文件中
def export_excel(store_filename, datas=None):
    print("--------export_excel--------")
    print "\n\n"

    if os.path.exists(store_filename):
        os.remove(store_filename)

    wb = Workbook()
    first_sheet = wb.active
    first_sheet.title = u"sheet1"

    if datas is not None:
        # print "datas: ", datas
        i = 1
        j = 1
        title_content = {}
        for k in range(len(title)):
            # print "title" + bytes(k) + ": ", title[k]
            if title[k] == 'files':
                first_sheet.cell(row=i, column=j, value="fileId")
                title_content[j] = "fileId"
                j += 1
                first_sheet.cell(row=i, column=j, value="fileName")
                title_content[j] = "fileName"
                j += 1
            else:
                first_sheet.cell(row=i, column=j, value=title[k])
                title_content[j] = title[k]
                j += 1

        first_sheet.cell(row=i, column=j, value='iscreated')
        title_content[j] = 'iscreated'

        # print title_content
        i += 1
        for data in datas:
            # print data
            for key in data:
                # if data[key] is None:
                #     break
                if key == 'files':
                    files = data[key]
                    file_ids = ""
                    file_names = ""
                    if files is not None:
                        for k in range(len(files)):
                            if k > 0:
                                file_ids = file_ids + "," + files[k]['id']
                                file_names = file_names + "," + files[k]['fileName']
                            else:
                                file_ids = files[k]['id']
                                file_names = files[k]['fileName']
                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('fileId')],
                                         value=file_ids)
                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('fileName')],
                                         value=file_names)
                else:
                    for title_num in title_content:
                        # print(title_num, title_content[title_num])
                        if key == title_content[title_num] and data[key] is not None:
                            first_sheet.cell(row=i, column=title_num, value=data[key])

            i += 1
    wb.save(store_filename)
    print "--------Excel save success--------"
    print "\n\n"


# 将数据导出phabricator类型的excel文件中
def export_excel_phabricator(store_filename, datas=None):
    print("Export Excel Phabricator")
    print "\n\n"

    if os.path.exists(store_filename):
        os.remove(store_filename)

    wb = Workbook()
    first_sheet = wb.active
    first_sheet.title = u"sheet1"

    if datas is not None:
        # print "datas: ", datas
        i = 1
        j = 1

        # write title
        title_content = {}
        for k in range(len(columnName)):
            print "columnName: num = " + bytes(k + 1) + ": ", columnName[k]
            first_sheet.cell(row=i, column=j, value=columnName[k])
            title_content[j] = columnName[k]
            j += 1
        # print title_content
        i += 1

        # write content
        for data in datas:
            print 'data: ', data
            if data['iscreated'] == util.IS_CREATED_TRUE:
                continue

            description = ''
            for key in data:
                # print "data: key = " + bytes(key) + ": ", data[key]
                if key == 'filename':
                    # print "filename: key = " + bytes(key) + ": ", data[key]
                    if data[key] is not None:
                        files = data[key].split(unicode(',', 'utf-8'))
                        file_paths = ''
                        if len(files) > 0:
                            num = 0
                            for file_path in files:
                                num += 1
                                if num != len(files):
                                    file_paths += util.attach_path + file_path + ','
                                else:
                                    file_paths += util.attach_path + file_path
                        else:
                            file_paths = util.attach_path + files

                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('File_location')],
                                         value=file_paths)
                elif key == 'topic':
                    # print "topic: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    files = unicode('【品质处】', 'utf-8') + files
                    if files is not None:
                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('Title')],
                                         value=files)
                elif key == 'grade':
                    # print "grade: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('Priority')],
                                         value=priority[files])
                elif key == 'estimateddate':
                    # print "estimatedDate: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        date = files.split(' ')
                        first_sheet.cell(row=i,
                                         column=list(title_content.keys())[
                                             list(title_content.values()).index('DueDate')],
                                         value=date[0])
                elif key == 'stage':
                    # print "stage: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        if 'S1' in files:
                            files = 'S1'
                        elif 'S2' in files:
                            files = 'S2'
                        elif 'S3' in files:
                            files = 'S3'
                        elif 'P' in files:
                            files = 'P'
                        if files is not None:
                            first_sheet.cell(row=i,
                                             column=list(title_content.keys())[
                                                 list(title_content.values()).index('HWVersion')],
                                             value=files)
                elif key == 'descr':
                    # print "descr: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        files = replace_invalid_content(files)
                        description = 'Description: ' + files + '\n' + description
                elif key == 'creatorname':
                    # print "creatorname: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        description += 'Creator_Name: ' + files + '\n'
                elif key == 'creatorid':
                    # print "creatorId: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        description += 'Creator_Id: ' + files + '\n'
                elif key == 'creatoremail':
                    print "creatoremail: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        description += 'Creator_Email: ' + files + '\n'
                elif key == 'creatortel':
                    # print "creatortel: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        description += 'Creator_Tel: ' + files + '\n'
                elif key == 'createtime':
                    # print "createtime: key = " + bytes(key) + ": ", data[key]
                    files = data[key]
                    if files is not None:
                        description += 'Creator_Time: ' + files + '\n'

            first_sheet.cell(row=i,
                             column=list(title_content.keys())[
                                 list(title_content.values()).index('Description')],
                             value=description)
            first_sheet.cell(row=i,
                             column=list(title_content.keys())[
                                 list(title_content.values()).index('Assigned')],
                             value=util.assign_username)
            first_sheet.cell(row=i,
                             column=list(title_content.keys())[
                                 list(title_content.values()).index('Visible')],
                             value=util.project_visible)

            i += 1

    wb.save(store_filename)
    print "--------Excel save success--------"
    print "\n\n"


# 导入excel文件中的数据
def load_excel(load_filename):
    print("loading Excel")
    # 打开一个workbook
    if not os.path.exists(load_filename):
        print "No corresponding Excel file found."

    try:
        wb = load_workbook(load_filename)
    except Exception, e:
        print "Excel is wrong : ", e
    else:
        # print(wb.sheetnames)
        sheetnames = wb.sheetnames
        first_sheet = wb[sheetnames[0]]
        print "Work Sheet Title : ", first_sheet.title
        # print "Work Sheet Max Rows:", first_sheet.max_row
        # print "Work Sheet Min Rows:", first_sheet.min_row
        # print "Work Sheet Max Cols:", first_sheet.max_column
        # print "Work Sheet Min Cols:", first_sheet.min_column
        post_data = []
        headers = []
        # print "min_row", first_sheet.min_row
        # print "max_row", first_sheet.max_row
        for row in range(first_sheet.min_row, first_sheet.max_row + 1):
            post_content = {}
            for column in range(first_sheet.min_column, first_sheet.max_column + 1):
                if row == 1:
                    value = first_sheet.cell(row=row, column=column).value
                    headers.append(value.lower().replace(' ', '').replace('\n', ''))
                else:
                    if column == 1 and not first_sheet.cell(row=row, column=column).value:
                        continue
                    else:
                        post_content[headers[column - 1]] = first_sheet.cell(row=row, column=column).value
            if post_content:
                print 'execl_content: ', post_content
                post_data.append(post_content)
        print "load excel : ", post_data
        return post_data


def replace_invalid_content(content=''):
    content = content.replace('<p>', '').replace('</p>', '').replace('&nbsp;', '')
    return content
